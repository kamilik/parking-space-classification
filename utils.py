"""
utils.py — Утилиты обучения для классификатора занятости парковочных мест.

Данный модуль предоставляет:

* ``EarlyStopping``      — остановка обучения при прекращении улучшения потерь на валидации.
* ``ModelCheckpoint``    — автоматическое сохранение лучших весов модели на диск.
* ``compute_metrics``    — точность, precision, recall, F1 и ROC-AUC по истинным
                           меткам и предсказанным вероятностям.
* ``plot_training_curves``  — графики потерь и точности по эпохам.
* ``plot_confusion_matrix`` — тепловая карта матрицы ошибок с аннотациями.
* ``plot_roc_curve``        — ROC-кривая с аннотацией значения AUC.
* ``save_comparison_table`` — сводная таблица CSV + XLSX по всем пяти архитектурам.
* ``analyze_results``       — текстовый анализ каждой модели с итоговой
                              рекомендацией.

Использование
-----
    from utils import (
        EarlyStopping, ModelCheckpoint, compute_metrics,
        plot_training_curves, plot_confusion_matrix, plot_roc_curve,
        save_comparison_table, analyze_results,
    )
"""

import logging
import math
from pathlib import Path
from typing import Any

import matplotlib
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import torch
import torch.nn as nn
from sklearn.metrics import (
    accuracy_score,
    auc,
    confusion_matrix,
    f1_score,
    precision_score,
    recall_score,
    roc_auc_score,
    roc_curve,
)

# Использовать неинтерактивный бэкенд, чтобы графики можно было сохранять
# в средах без дисплея (Google Colab, серверы).
matplotlib.use("Agg")

# ---------------------------------------------------------------------------
# Логгер уровня модуля
# ---------------------------------------------------------------------------
logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Ранняя остановка
# ---------------------------------------------------------------------------
class EarlyStopping:
    """
    Отслеживать валидационную метрику и сигнализировать о необходимости досрочной остановки.

    По умолчанию (``mode="max"``), обучение считается остановившимся, если
    отслеживаемая метрика (например, F1-score) не улучшается как минимум на
    ``min_delta`` в течение ``patience`` последовательных эпох.

    Параметры
    ----------
    patience:
        Количество эпох без улучшения, допустимых перед остановкой.
    min_delta:
        Минимальное абсолютное улучшение, необходимое для сброса счётчика ожидания.
        По умолчанию ``0.0``: любое строгое улучшение сбрасывает счётчик.
    mode:
        ``"max"`` — чем больше, тем лучше (F1, accuracy).
        ``"min"`` — чем меньше, тем лучше (loss).

    Атрибуты
    ----------
    best_score:
        Лучшее наблюдённое значение метрики на данный момент.
    counter:
        Текущее число последовательных эпох без улучшения.
    """

    def __init__(self, patience: int, min_delta: float = 0.0, mode: str = "max") -> None:
        if patience < 1:
            raise ValueError(f"patience must be >= 1, got {patience}")
        if mode not in ("min", "max"):
            raise ValueError(f"mode must be 'min' or 'max', got '{mode}'")
        self.patience: int = patience
        self.min_delta: float = min_delta
        self.mode: str = mode
        self.best_score: float = -math.inf if mode == "max" else math.inf
        self.counter: int = 0

    def _is_improvement(self, current: float) -> bool:
        if self.mode == "max":
            return current > self.best_score + self.min_delta
        return current < self.best_score - self.min_delta

    # ------------------------------------------------------------------
    def __call__(self, metric_value: float) -> bool:
        """
        Оценить последнее значение метрики и решить, следует ли останавливаться.

        Параметры
        ----------
        metric_value:
            Валидационная метрика, измеренная в конце текущей эпохи
            (например, F1-score при ``mode="max"``).

        Возвращает
        -------
        bool
            ``True``, когда лимит ожидания исчерпан и обучение следует
            прекратить; ``False`` в противном случае.
        """
        if self._is_improvement(metric_value):
            logger.debug(
                "EarlyStopping: metric improved %.6f → %.6f",
                self.best_score,
                metric_value,
            )
            self.best_score = metric_value
            self.counter = 0
            return False

        self.counter += 1
        logger.debug(
            "EarlyStopping: no improvement (%.6f vs best %.6f), counter=%d/%d",
            metric_value,
            self.best_score,
            self.counter,
            self.patience,
        )
        if self.counter >= self.patience:
            logger.info(
                "EarlyStopping triggered after %d epochs with no improvement.",
                self.patience,
            )
            return True
        return False

    # ------------------------------------------------------------------
    def __repr__(self) -> str:
        return (
            f"{self.__class__.__name__}("
            f"patience={self.patience}, "
            f"min_delta={self.min_delta}, "
            f"mode='{self.mode}', "
            f"best_score={self.best_score:.6f}, "
            f"counter={self.counter})"
        )


# ---------------------------------------------------------------------------
# Сохранение контрольных точек модели
# ---------------------------------------------------------------------------
class ModelCheckpoint:
    """
    Сохранять веса модели каждый раз, когда отслеживаемая метрика достигает нового максимума.

    По умолчанию (``mode="max"``), сохранение происходит, когда метрика (например, F1-score)
    превышает предыдущий лучший результат.

    Параметры
    ----------
    save_path:
        Полный путь файловой системы (включая имя файла), куда будет записан файл ``.pth``.
        Родительские директории создаются автоматически.
    mode:
        ``"max"`` — чем больше, тем лучше (F1, accuracy).
        ``"min"`` — чем меньше, тем лучше (loss).

    Атрибуты
    ----------
    best_value:
        Лучшее значение метрики, для которого был сохранён чекпоинт.
    """

    def __init__(self, save_path: Path, mode: str = "max") -> None:
        if mode not in ("min", "max"):
            raise ValueError(f"mode must be 'min' or 'max', got '{mode}'")
        self.save_path: Path = Path(save_path)
        self.save_path.parent.mkdir(parents=True, exist_ok=True)
        self.mode: str = mode
        self.best_value: float = -math.inf if mode == "max" else math.inf

    def _is_improvement(self, current: float) -> bool:
        if self.mode == "max":
            return current > self.best_value
        return current < self.best_value

    # ------------------------------------------------------------------
    def __call__(self, metric_value: float, model: nn.Module) -> None:
        """
        Сохранить веса модели, если метрика улучшилась.

        Параметры
        ----------
        metric_value:
            Валидационная метрика (например, F1-score) в конце текущей эпохи.
        model:
            Модель PyTorch, state dict которой следует сохранить.
        """
        if self._is_improvement(metric_value):
            logger.info(
                "ModelCheckpoint: metric improved %.6f → %.6f — saving to %s",
                self.best_value,
                metric_value,
                self.save_path,
            )
            self.best_value = metric_value
            torch.save(model.state_dict(), self.save_path)
        else:
            logger.debug(
                "ModelCheckpoint: metric %.6f did not improve best %.6f — skip",
                metric_value,
                self.best_value,
            )

    # ------------------------------------------------------------------
    def __repr__(self) -> str:
        return (
            f"{self.__class__.__name__}("
            f"save_path={self.save_path!r}, "
            f"mode='{self.mode}', "
            f"best_value={self.best_value:.6f})"
        )


# ---------------------------------------------------------------------------
# Метрики
# ---------------------------------------------------------------------------
def compute_metrics(
    y_true: np.ndarray,
    y_pred: np.ndarray,
    y_proba: np.ndarray,
) -> dict[str, float]:
    """
    Вычислить метрики бинарной классификации по истинным меткам и предсказаниям.

    Параметры
    ----------
    y_true:
        Одномерный целочисленный массив истинных индексов классов (0 = Empty,
        1 = Occupied).
    y_pred:
        Одномерный целочисленный массив предсказанных индексов классов (argmax логитов).
    y_proba:
        Одномерный массив вещественных чисел — предсказанные вероятности для
        *положительного* класса (индекс 1 — Occupied). Используется исключительно
        для вычисления ROC-AUC.

    Возвращает
    -------
    dict[str, float]
        Словарь со следующими ключами:

        * ``accuracy``  — доля правильно классифицированных образцов.
        * ``precision`` — взвешенная точность по обоим классам.
        * ``recall``    — взвешенная полнота по обоим классам.
        * ``f1``        — взвешенный F1-score.
        * ``roc_auc``   — площадь под ROC-кривой.

    Исключения
    ------
    ValueError
        Если входные массивы имеют несовпадающую длину.
    """
    if len(y_true) != len(y_pred) or len(y_true) != len(y_proba):
        raise ValueError(
            "y_true, y_pred, and y_proba must all have the same length; "
            f"got {len(y_true)}, {len(y_pred)}, {len(y_proba)}"
        )

    accuracy: float = float(accuracy_score(y_true, y_pred))
    precision: float = float(
        precision_score(y_true, y_pred, average="weighted", zero_division=0)
    )
    recall: float = float(
        recall_score(y_true, y_pred, average="weighted", zero_division=0)
    )
    f1: float = float(
        f1_score(y_true, y_pred, average="weighted", zero_division=0)
    )
    roc_auc: float = float(roc_auc_score(y_true, y_proba))

    metrics: dict[str, float] = {
        "accuracy": accuracy,
        "precision": precision,
        "recall": recall,
        "f1": f1,
        "roc_auc": roc_auc,
    }

    logger.info(
        "Metrics — acc=%.4f | prec=%.4f | rec=%.4f | f1=%.4f | auc=%.4f",
        accuracy,
        precision,
        recall,
        f1,
        roc_auc,
    )
    return metrics


# ---------------------------------------------------------------------------
# Вспомогательные функции построения графиков
# ---------------------------------------------------------------------------
def plot_training_curves(
    history: dict[str, list[float]],
    model_name: str,
    save_dir: Path,
) -> None:
    """
    Построить и сохранить графики кривых обучения для одной модели.

    Записываются два отдельных файла PNG:

    * ``{model_name}_loss.png``     — потери на обучении и валидации по эпохам.
    * ``{model_name}_accuracy.png`` — точность на обучении и валидации по эпохам.

    Параметры
    ----------
    history:
        Словарь, сформированный в процессе обучения, с ключами ``train_loss``,
        ``val_loss``, ``train_acc`` и ``val_acc``, каждый из которых содержит
        список скалярных значений — по одному на эпоху.
    model_name:
        Читаемый идентификатор модели, используемый в заголовке графика и
        имени файла (например, ``"ResNet18"``).
    save_dir:
        Директория, в которую будут сохранены файлы PNG. Создаётся автоматически,
        если не существует.
    """
    save_dir = Path(save_dir)
    save_dir.mkdir(parents=True, exist_ok=True)

    epochs: list[int] = list(range(1, len(history["train_loss"]) + 1))

    # ── График потерь ────────────────────────────────────────────────────────
    fig_loss, ax_loss = plt.subplots(figsize=(8, 5))
    ax_loss.plot(epochs, history["train_loss"], label="Train Loss", linewidth=2)
    ax_loss.plot(epochs, history["val_loss"], label="Val Loss", linewidth=2)
    ax_loss.set_title(f"{model_name} — Training & Validation Loss", fontsize=14)
    ax_loss.set_xlabel("Epoch", fontsize=12)
    ax_loss.set_ylabel("Loss", fontsize=12)
    ax_loss.legend(fontsize=11)
    ax_loss.grid(True, linestyle="--", alpha=0.6)
    fig_loss.tight_layout()
    loss_path: Path = save_dir / f"{model_name}_loss.png"
    fig_loss.savefig(loss_path, dpi=150, bbox_inches="tight")
    plt.close(fig_loss)
    logger.info("Saved loss curve → %s", loss_path)

    # ── График точности ────────────────────────────────────────────────────────
    fig_acc, ax_acc = plt.subplots(figsize=(8, 5))
    ax_acc.plot(epochs, history["train_acc"], label="Train Accuracy", linewidth=2)
    ax_acc.plot(epochs, history["val_acc"], label="Val Accuracy", linewidth=2)
    ax_acc.set_title(f"{model_name} — Training & Validation Accuracy", fontsize=14)
    ax_acc.set_xlabel("Epoch", fontsize=12)
    ax_acc.set_ylabel("Accuracy", fontsize=12)
    ax_acc.legend(fontsize=11)
    ax_acc.grid(True, linestyle="--", alpha=0.6)
    fig_acc.tight_layout()
    acc_path: Path = save_dir / f"{model_name}_accuracy.png"
    fig_acc.savefig(acc_path, dpi=150, bbox_inches="tight")
    plt.close(fig_acc)
    logger.info("Saved accuracy curve → %s", acc_path)


# ---------------------------------------------------------------------------
def plot_confusion_matrix(
    y_true: np.ndarray,
    y_pred: np.ndarray,
    class_names: tuple[str, ...] | list[str],
    model_name: str,
    save_dir: Path,
) -> None:
    """
    Построить и сохранить аннотированную матрицу ошибок для одной модели.

    Цветовая карта варьируется от белого (нулевые значения) до насыщенного синего
    (максимальное значение). Каждая ячейка аннотирована целочисленным значением.

    Параметры
    ----------
    y_true:
        Одномерный целочисленный массив истинных индексов классов.
    y_pred:
        Одномерный целочисленный массив предсказанных индексов классов.
    class_names:
        Упорядоченная последовательность меток классов, соответствующая порядку
        индексов в ``y_true`` и ``y_pred`` (например, ``("Empty", "Occupied")``).
    model_name:
        Читаемый идентификатор модели, используемый в заголовке и имени файла.
    save_dir:
        Директория, в которую будет сохранён файл PNG. Создаётся автоматически,
        если не существует.
    """
    save_dir = Path(save_dir)
    save_dir.mkdir(parents=True, exist_ok=True)

    cm: np.ndarray = confusion_matrix(y_true, y_pred)
    num_classes: int = len(class_names)

    fig, ax = plt.subplots(figsize=(6, 5))
    im = ax.imshow(cm, interpolation="nearest", cmap="Blues")
    fig.colorbar(im, ax=ax)

    ax.set_title(f"{model_name} — Confusion Matrix", fontsize=14)
    ax.set_xlabel("Predicted Label", fontsize=12)
    ax.set_ylabel("True Label", fontsize=12)

    tick_marks: list[int] = list(range(num_classes))
    ax.set_xticks(tick_marks)
    ax.set_yticks(tick_marks)
    ax.set_xticklabels(class_names, fontsize=11)
    ax.set_yticklabels(class_names, fontsize=11)

    # Аннотировать каждую ячейку её значением; использовать белый текст на тёмных ячейках.
    thresh: float = cm.max() / 2.0
    for row in range(num_classes):
        for col in range(num_classes):
            text_color: str = "white" if cm[row, col] > thresh else "black"
            ax.text(
                col,
                row,
                str(cm[row, col]),
                ha="center",
                va="center",
                color=text_color,
                fontsize=13,
                fontweight="bold",
            )

    fig.tight_layout()
    cm_path: Path = save_dir / f"{model_name}_confusion_matrix.png"
    fig.savefig(cm_path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    logger.info("Saved confusion matrix → %s", cm_path)


# ---------------------------------------------------------------------------
def plot_roc_curve(
    y_true: np.ndarray,
    y_proba: np.ndarray,
    model_name: str,
    save_dir: Path,
) -> None:
    """
    Построить и сохранить ROC-кривую со значением AUC в легенде.

    Диагональная референсная линия (базовый уровень случайного классификатора)
    отображается серым цветом.

    Параметры
    ----------
    y_true:
        Одномерный целочисленный массив истинных индексов классов.
    y_proba:
        Одномерный массив вещественных чисел — предсказанные вероятности для
        положительного класса (индекс 1 — Occupied).
    model_name:
        Читаемый идентификатор модели, используемый в заголовке и имени файла.
    save_dir:
        Директория, в которую будет сохранён файл PNG. Создаётся автоматически,
        если не существует.
    """
    save_dir = Path(save_dir)
    save_dir.mkdir(parents=True, exist_ok=True)

    fpr: np.ndarray
    tpr: np.ndarray
    _thresholds: np.ndarray
    fpr, tpr, _thresholds = roc_curve(y_true, y_proba)
    roc_auc_value: float = float(auc(fpr, tpr))

    fig, ax = plt.subplots(figsize=(7, 5))

    # Диагональная референсная линия (случайный классификатор).
    ax.plot(
        [0.0, 1.0],
        [0.0, 1.0],
        color="grey",
        linestyle="--",
        linewidth=1.5,
        label="Random classifier (AUC = 0.50)",
    )

    # ROC-кривая данной модели.
    ax.plot(
        fpr,
        tpr,
        color="steelblue",
        linewidth=2.5,
        label=f"{model_name} (AUC = {roc_auc_value:.4f})",
    )

    ax.set_title(f"{model_name} — ROC Curve", fontsize=14)
    ax.set_xlabel("False Positive Rate", fontsize=12)
    ax.set_ylabel("True Positive Rate", fontsize=12)
    ax.legend(fontsize=11, loc="lower right")
    ax.grid(True, linestyle="--", alpha=0.6)
    ax.set_xlim([0.0, 1.0])
    ax.set_ylim([0.0, 1.05])

    fig.tight_layout()
    roc_path: Path = save_dir / f"{model_name}_roc_curve.png"
    fig.savefig(roc_path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    logger.info("Saved ROC curve → %s", roc_path)


# ---------------------------------------------------------------------------
# Сравнительная таблица
# ---------------------------------------------------------------------------
def save_comparison_table(
    results: list[dict[str, Any]],
    save_dir: Path,
) -> None:
    """
    Собрать, аннотировать и сохранить сравнительную таблицу по всем архитектурам.

    Таблица сортируется по F1-score в порядке убывания. Добавляется столбец ``Best``,
    значение которого равно ``"*"`` для лучшей модели и пусто для всех остальных.

    В директории ``save_dir`` записываются два файла:

    * ``comparison.csv``  — разделённый запятыми, UTF-8.
    * ``comparison.xlsx`` — книга Excel с автоматически подобранной шириной столбцов
      и строкой лучшей модели, выделенной светло-зелёным цветом.

    Параметры
    ----------
    results:
        Список словарей — по одному на архитектуру — каждый из которых содержит
        следующие ключи (все должны присутствовать):

        * ``Architecture``   — строка с названием модели.
        * ``Accuracy``       — вещественное число в [0, 1].
        * ``Precision``      — вещественное число в [0, 1].
        * ``Recall``         — вещественное число в [0, 1].
        * ``F1``             — вещественное число в [0, 1].
        * ``ROC_AUC``        — вещественное число в [0, 1].
        * ``Inference_Time`` — секунды на батч (вещественное число).
        * ``Parameters``     — общее количество обучаемых параметров (int).
        * ``Model_Size``     — размер чекпоинта в мегабайтах (вещественное число).
        * ``Epochs``         — фактически пройденное число эпох (int).
        * ``Training_Time``  — общее астрономическое время обучения в секундах (вещественное число).

    save_dir:
        Директория, в которую будут записаны выходные файлы. Создаётся
        автоматически, если не существует.
    """
    save_dir = Path(save_dir)
    save_dir.mkdir(parents=True, exist_ok=True)

    # Определить канонический порядок столбцов.
    columns: list[str] = [
        "Architecture",
        "Accuracy",
        "Precision",
        "Recall",
        "F1",
        "ROC AUC",
        "Training Time",
        "Inference Time (ms/image)",
        "FPS",
        "Model Size (MB)",
        "Number of Parameters",
        "Epochs",
    ]

    df: pd.DataFrame = pd.DataFrame(results, columns=columns)
    df.sort_values(by="F1", ascending=False, inplace=True)
    df.reset_index(drop=True, inplace=True)

    # Отметить лучшую модель (после сортировки первая строка — наивысший F1).
    df.insert(0, "Best", "")
    df.at[0, "Best"] = "*"

    # ── CSV ──────────────────────────────────────────────────────────────────
    csv_path: Path = save_dir / "comparison.csv"
    df.to_csv(csv_path, index=False, encoding="utf-8")
    logger.info("Saved comparison CSV → %s", csv_path)

    # ── XLSX ─────────────────────────────────────────────────────────────────
    xlsx_path: Path = save_dir / "comparison.xlsx"
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Comparison")

        workbook = writer.book
        worksheet = writer.sheets["Comparison"]

        # Автоматически подобрать ширину столбцов.
        for col_cells in worksheet.columns:
            max_length: int = 0
            col_letter: str = col_cells[0].column_letter
            for cell in col_cells:
                if cell.value is not None:
                    cell_str_len: int = len(str(cell.value))
                    if cell_str_len > max_length:
                        max_length = cell_str_len
            adjusted_width: float = max_length + 4
            worksheet.column_dimensions[col_letter].width = adjusted_width

        # Выделить строку лучшей модели (строка 2 — строка 1 является заголовком).
        from openpyxl.styles import PatternFill

        green_fill = PatternFill(
            start_color="C6EFCE",
            end_color="C6EFCE",
            fill_type="solid",
        )
        num_cols: int = len(df.columns)
        for col_idx in range(1, num_cols + 1):
            worksheet.cell(row=2, column=col_idx).fill = green_fill

    logger.info("Saved comparison XLSX → %s", xlsx_path)


# ---------------------------------------------------------------------------
# Текстовый анализ
# ---------------------------------------------------------------------------

# Текстовые описания по архитектурам, заполняемые фактическими значениями метрик
# из списка результатов во время вызова.
_ARCH_PROFILES: dict[str, dict[str, str]] = {
    "ResNet18": {
        "strengths": (
            "Очень быстрое обучение и инференс благодаря неглубокой остаточной "
            "сети (18 слоёв). Малое количество параметров упрощает развёртывание "
            "и дообучение на ограниченном оборудовании. Остаточные связи устраняют "
            "проблему затухающих градиентов и обеспечивают стабильное обучение с первой эпохи."
        ),
        "weaknesses": (
            "Меньшая ёмкость по сравнению с более глубокими или широкими сетями; "
            "может выйти на плато с более низким потолком точности для сложных "
            "визуальных паттернов. Хуже представляет тонкие вариации освещённости "
            "или окклюзии по сравнению с архитектурами EfficientNet или ViT."
        ),
        "speed_note": (
            "Самая быстрая архитектура в сравнении. Подходит для инференса "
            "в реальном времени на граничных устройствах или при ограниченных ресурсах."
        ),
    },
    "DenseNet121": {
        "strengths": (
            "Плотная связность стимулирует повторное использование признаков между "
            "слоями, что может улучшить обобщение при ограниченных данных. Сильный "
            "поток градиентов через сеть снижает переобучение на относительно "
            "небольших датасетах, таких как PKLot."
        ),
        "weaknesses": (
            "Больший объём используемой памяти при обучении по сравнению с ResNet18, "
            "так как активации всех предыдущих слоёв должны конкатенироваться. "
            "Инференс медленнее, чем у лёгких моделей вроде MobileNetV3."
        ),
        "speed_note": (
            "Умеренная скорость. Потребление памяти при обучении выше среднего; "
            "задержка инференса приемлема для пакетной обработки."
        ),
    },
    "EfficientNet-B0": {
        "strengths": (
            "Благодаря составному масштабированию глубины, ширины и разрешения "
            "EfficientNet-B0 обеспечивает высокую точность при небольшом бюджете "
            "параметров. Глубинно-разделимые свёртки делают её вычислительно "
            "эффективной при сохранении высокой представительной мощи."
        ),
        "weaknesses": (
            "Несколько более сложная динамика обучения по сравнению с обычными "
            "свёрточными сетями; расписание скорости обучения может быть более "
            "чувствительным. Требует тщательной аугментации данных во избежание "
            "переобучения на небольших датасетах."
        ),
        "speed_note": (
            "Хороший баланс точности и скорости. Одна из лучших архитектур "
            "по соотношению точности к числу FLOP в данном сравнении."
        ),
    },
    "MobileNetV3-Large": {
        "strengths": (
            "Специально разработана для инференса на устройстве: инвертированные "
            "остатки и активации hard-swish минимизируют задержку. Наименьший "
            "размер модели и наибыстрейший инференс в группе, что делает её "
            "идеальной для встраиваемых или мобильных развёртываний."
        ),
        "weaknesses": (
            "Более низкая представительная ёмкость по сравнению с EfficientNet "
            "или DenseNet для тонкозернистой классификации. Может потребоваться "
            "больше аугментаций или более длительное обучение для достижения "
            "результатов более глубоких архитектур."
        ),
        "speed_note": (
            "Самый быстрый или совместно самый быстрый инференс. Рекомендуется "
            "при строгих ограничениях на задержку или размер модели."
        ),
    },
    "ViT-B/16": {
        "strengths": (
            "Vision Transformer улавливает дальнодействующие пространственные "
            "зависимости через многоголовое самовнимание, что может быть "
            "преимуществом, когда признаки занятости парковки распределены по "
            "изображению (например, кузов автомобиля виден только в одном углу). "
            "Предобучение на больших корпусах (ImageNet-21k) переносит богатые "
            "представления признаков."
        ),
        "weaknesses": (
            "Требует наибольшего количества параметров и наибольшего времени "
            "обучения в сравнении. Архитектуры трансформеров, как правило, "
            "нуждаются в больших датасетах для раскрытия полного потенциала; "
            "при работе только с PKLot перенос обучения снижает, но не устраняет "
            "этот риск. Инференс медленнее, чем у всех CNN-аналогов."
        ),
        "speed_note": (
            "Самая медленная архитектура в сравнении. Лучше всего подходит "
            "для офлайн пакетного инференса, где точность важнее задержки."
        ),
    },
}

# Резервный профиль для архитектур, не перечисленных выше.
_DEFAULT_PROFILE: dict[str, str] = {
    "strengths": "Перенос обучения с весов ImageNet обеспечивает сильную базу признаков.",
    "weaknesses": "Профиль недоступен; оценивайте эмпирически по таблице метрик.",
    "speed_note": "Характеристики скорости не профилированы; измерьте по столбцу Inference_Time.",
}


def analyze_results(results: list[dict[str, Any]]) -> str:
    """
    Сформировать исчерпывающий текстовый анализ всех обученных моделей.

    Для каждой архитектуры анализ включает:

    * **Сильные стороны** — архитектурные преимущества.
    * **Слабые стороны** — известные ограничения.
    * **Скорость** — качественные и количественные замечания на основе фактических
      измерений ``Inference_Time`` и ``Training_Time``.
    * **Точность** — количественная сводка по фактическим метрикам.
    * **Вывод** — итоговый вердикт по каждой модели.

    После анализа по каждой модели итоговый раздел выбирает лучшую модель по
    наибольшему F1-score и объясняет логику выбора.

    Параметры
    ----------
    results:
        Список словарей с той же схемой, что принимает :func:`save_comparison_table`.

    Возвращает
    -------
    str
        Многосекционный текст анализа, готовый для вывода, логирования или
        встраивания в ячейку Markdown Jupyter-ноутбука.
    """
    if not results:
        return "No results to analyse."

    # Сортировать по F1 в порядке убывания для определения рейтинга.
    sorted_results: list[dict[str, Any]] = sorted(
        results, key=lambda r: r.get("F1", 0.0), reverse=True
    )
    best: dict[str, Any] = sorted_results[0]
    best_arch: str = best.get("Architecture", "Unknown")

    lines: list[str] = []

    lines.append("=" * 72)
    lines.append("  MODEL ANALYSIS — PARKING SPOT OCCUPANCY CLASSIFICATION")
    lines.append("=" * 72)
    lines.append("")

    # ── Блоки по каждой модели ────────────────────────────────────────────────
    for rank, res in enumerate(sorted_results, start=1):
        arch: str = res.get("Architecture", "Unknown")
        acc: float = res.get("Accuracy", 0.0)
        prec: float = res.get("Precision", 0.0)
        rec: float = res.get("Recall", 0.0)
        f1: float = res.get("F1", 0.0)
        roc: float = res.get("ROC AUC", 0.0)
        inf_time_ms: float = res.get("Inference Time (ms/image)", 0.0)
        fps: float = res.get("FPS", 0.0)
        params: int = res.get("Number of Parameters", 0)
        model_mb: float = res.get("Model Size (MB)", 0.0)
        epochs: int = res.get("Epochs", 0)
        train_time: float = res.get("Training Time", 0.0)

        profile: dict[str, str] = _ARCH_PROFILES.get(arch, _DEFAULT_PROFILE)

        header: str = f"  #{rank}  {arch}"
        if arch == best_arch:
            header += "  [BEST MODEL]"
        lines.append("-" * 72)
        lines.append(header)
        lines.append("-" * 72)

        lines.append("")
        lines.append("  Metrics")
        lines.append(f"    Accuracy      : {acc:.4f}  ({acc * 100:.2f} %)")
        lines.append(f"    Precision     : {prec:.4f}  ({prec * 100:.2f} %)")
        lines.append(f"    Recall        : {rec:.4f}  ({rec * 100:.2f} %)")
        lines.append(f"    F1 Score      : {f1:.4f}  ({f1 * 100:.2f} %)")
        lines.append(f"    ROC-AUC       : {roc:.4f}")
        lines.append("")
        lines.append("  Runtime")
        lines.append(
            f"    Epochs trained: {epochs}"
        )
        lines.append(
            f"    Training time : {train_time:.1f} s  "
            f"({train_time / 60:.1f} min)"
        )
        lines.append(
            f"    Inference time: {inf_time_ms:.2f} ms / image"
        )
        lines.append(
            f"    FPS           : {fps:.1f}"
        )
        lines.append(
            f"    Parameters    : {params:,}"
        )
        lines.append(
            f"    Model size    : {model_mb:.2f} MB"
        )
        lines.append("")
        lines.append("  Strengths")
        for sentence in _wrap_text(profile["strengths"], width=68, indent="    "):
            lines.append(sentence)
        lines.append("")
        lines.append("  Weaknesses")
        for sentence in _wrap_text(profile["weaknesses"], width=68, indent="    "):
            lines.append(sentence)
        lines.append("")
        lines.append("  Speed")
        for sentence in _wrap_text(profile["speed_note"], width=68, indent="    "):
            lines.append(sentence)
        lines.append("")

        # Вердикт по каждой модели.
        if f1 >= 0.98:
            verdict = "Outstanding classification performance. Highly recommended."
        elif f1 >= 0.95:
            verdict = "Excellent performance. Suitable for production use."
        elif f1 >= 0.90:
            verdict = "Good performance. May benefit from additional fine-tuning."
        elif f1 >= 0.80:
            verdict = "Acceptable performance. Further hyperparameter search advised."
        else:
            verdict = "Below expectations. Consider deeper fine-tuning or more data."

        lines.append("  Conclusion")
        lines.append(f"    {verdict}")
        lines.append("")

    # ── Рекомендация лучшей модели ────────────────────────────────────────────
    lines.append("=" * 72)
    lines.append("  BEST MODEL RECOMMENDATION")
    lines.append("=" * 72)
    lines.append("")
    lines.append(f"  Selected architecture : {best_arch}")
    lines.append(f"  F1 Score              : {best.get('F1', 0.0):.4f}")
    lines.append(f"  Accuracy              : {best.get('Accuracy', 0.0):.4f}")
    lines.append(f"  ROC-AUC               : {best.get('ROC AUC', 0.0):.4f}")
    lines.append("")

    rationale_lines: list[str] = [
        f"{best_arch} was selected as the best model because it achieved the "
        "highest F1 score on the hold-out test set.  F1 is the primary "
        "selection criterion because it harmonises precision and recall, "
        "which is important for parking-occupancy detection: a false "
        "negative (reporting an occupied space as empty) wastes drivers' "
        "time, while a false positive (reporting an empty space as occupied) "
        "can misdirect navigation.  Maximising F1 ensures both error types "
        "are kept in check simultaneously.",
    ]
    best_profile: dict[str, str] = _ARCH_PROFILES.get(best_arch, _DEFAULT_PROFILE)
    rationale_lines.append(
        f"From an architectural standpoint, {best_arch} benefits from: "
        + best_profile["strengths"]
    )
    rationale_lines.append(
        "Given these properties and the measured metrics, "
        f"{best_arch} is recommended for deployment in the parking-space "
        "occupancy classifier system."
    )

    for para in rationale_lines:
        for wrapped_line in _wrap_text(para, width=68, indent="  "):
            lines.append(wrapped_line)
        lines.append("")

    lines.append("=" * 72)

    analysis_text: str = "\n".join(lines)
    logger.info("Analysis generated for %d models. Best: %s", len(results), best_arch)
    return analysis_text


# ---------------------------------------------------------------------------
# Внутренняя вспомогательная функция
# ---------------------------------------------------------------------------
def _wrap_text(text: str, width: int, indent: str = "") -> list[str]:
    """
    Перенести ``text`` по ``width`` символов на строку, добавляя префикс ``indent``.

    Параметры
    ----------
    text:
        Исходный текст для переноса.
    width:
        Максимальная ширина каждой выходной строки в символах (без учёта отступа).
    indent:
        Строка, добавляемая в начало каждой выходной строки.

    Возвращает
    -------
    list[str]
        Список перенесённых строк с применённым отступом.
    """
    import textwrap

    wrapped: list[str] = textwrap.wrap(text, width=width)
    return [indent + line for line in wrapped] if wrapped else [indent + text]
